from pathlib import Path

from openpyxl import Workbook, load_workbook


def _resolve_workbook_path(file_path: str):
    if not file_path:
        raise ValueError("file_path is required.")

    target = Path(file_path).expanduser()
    if target.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported.")
    return target


def _load_existing_workbook(file_path: str):
    target = _resolve_workbook_path(file_path)
    if not target.exists():
        raise FileNotFoundError(f"Workbook does not exist: {target}")
    workbook = load_workbook(target)
    return workbook, target


def _get_sheet(workbook, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return workbook[sheet_name]
    return workbook.active


def create_excel_file(file_path: str, sheet_name: str = "Sheet1"):
    target = _resolve_workbook_path(file_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name or "Sheet1"
    workbook.save(target)
    return f"Created workbook: {target} with sheet: {sheet.title}"


def list_workbook_sheets(file_path: str):
    workbook, target = _load_existing_workbook(file_path)
    return f"Workbook: {target}\nSheets: {', '.join(workbook.sheetnames)}"


def create_sheet(file_path: str, sheet_name: str):
    if not sheet_name:
        raise ValueError("sheet_name is required.")

    workbook, target = _load_existing_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        raise ValueError(f"Sheet already exists: {sheet_name}")

    workbook.create_sheet(title=sheet_name)
    workbook.save(target)
    return f"Created sheet '{sheet_name}' in workbook: {target}"


def delete_sheet(file_path: str, sheet_name: str):
    if not sheet_name:
        raise ValueError("sheet_name is required.")

    workbook, target = _load_existing_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    if len(workbook.sheetnames) == 1:
        raise ValueError("Cannot delete the only remaining sheet in the workbook.")

    del workbook[sheet_name]
    workbook.save(target)
    return f"Deleted sheet '{sheet_name}' from workbook: {target}"


def read_excel_range(file_path: str, sheet_name: str = "", cell_range: str = "A1:C10"):
    workbook, target = _load_existing_workbook(file_path)
    sheet = _get_sheet(workbook, sheet_name or None)
    values = []
    for row in sheet[cell_range]:
        values.append([cell.value for cell in row])

    return (
        f"Workbook: {target}\n"
        f"Sheet: {sheet.title}\n"
        f"Range: {cell_range}\n"
        f"Values: {values}"
    )


def write_excel_cells(file_path: str, sheet_name: str = "", cells: list | None = None):
    if not cells:
        raise ValueError("cells is required and must contain at least one cell update.")

    workbook, target = _load_existing_workbook(file_path)
    sheet = _get_sheet(workbook, sheet_name or None)

    written_cells = []
    for item in cells:
        if not isinstance(item, dict):
            raise ValueError("Each cells item must be an object like {'cell': 'A1', 'value': 'text'}.")

        cell_ref = item.get("cell")
        if not cell_ref:
            raise ValueError("Each cells item must include a cell field.")

        sheet[cell_ref] = item.get("value")
        written_cells.append(cell_ref)

    workbook.save(target)
    return f"Wrote {len(written_cells)} cells to {target}, sheet {sheet.title}: {', '.join(written_cells)}"


def append_excel_row(file_path: str, sheet_name: str = "", values: list | None = None):
    if values is None:
        raise ValueError("values is required.")

    workbook, target = _load_existing_workbook(file_path)
    sheet = _get_sheet(workbook, sheet_name or None)
    sheet.append(values)
    appended_row = sheet.max_row
    workbook.save(target)
    return f"Appended row {appended_row} to {target}, sheet {sheet.title}: {values}"


def find_in_excel(file_path: str, sheet_name: str = "", query: str = ""):
    if not query:
        raise ValueError("query is required.")

    workbook, target = _load_existing_workbook(file_path)
    sheets = [workbook[sheet_name]] if sheet_name else [workbook[name] for name in workbook.sheetnames]
    matches = []

    for sheet in sheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                cell_text = str(cell.value)
                if query.lower() in cell_text.lower():
                    matches.append({
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "value": cell_text,
                    })

    return (
        f"Workbook: {target}\n"
        f"Query: {query}\n"
        f"Matches: {matches}"
    )


def register(agent):
    agent.add_skill(
        name="create_excel_file",
        func=create_excel_file,
        description="Create a new .xlsx workbook with an optional initial sheet name.",
        parameters={
            "file_path": "string",
            "sheet_name": "string",
        },
    )
    agent.add_skill(
        name="list_workbook_sheets",
        func=list_workbook_sheets,
        description="List all sheet names in an existing .xlsx workbook.",
        parameters={
            "file_path": "string",
        },
    )
    agent.add_skill(
        name="create_sheet",
        func=create_sheet,
        description="Create a new sheet in an existing .xlsx workbook.",
        parameters={
            "file_path": "string",
            "sheet_name": "string",
        },
    )
    agent.add_skill(
        name="delete_sheet",
        func=delete_sheet,
        description="Delete a sheet from an existing .xlsx workbook.",
        parameters={
            "file_path": "string",
            "sheet_name": "string",
        },
    )
    agent.add_skill(
        name="read_excel_range",
        func=read_excel_range,
        description="Read a cell range from an .xlsx workbook sheet.",
        parameters={
            "file_path": "string",
            "sheet_name": "string",
            "cell_range": "string",
        },
    )
    agent.add_skill(
        name="write_excel_cells",
        func=write_excel_cells,
        description="Write one or more cell values into an .xlsx workbook sheet.",
        parameters={
            "file_path": "string",
            "sheet_name": "string",
            "cells": "array",
        },
    )
    agent.add_skill(
        name="append_excel_row",
        func=append_excel_row,
        description="Append a row of values to an .xlsx workbook sheet.",
        parameters={
            "file_path": "string",
            "sheet_name": "string",
            "values": "array",
        },
    )
    agent.add_skill(
        name="find_in_excel",
        func=find_in_excel,
        description="Find cells containing a query string in one sheet or across the entire workbook.",
        parameters={
            "file_path": "string",
            "sheet_name": "string",
            "query": "string",
        },
    )